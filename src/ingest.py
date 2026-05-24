"""
Load the source Excel sheets and enumerate the PDF corpus.

The hackathon ships everything in one xlsx (Business Rules, Submissions targets,
worked-example Reference row, 440-row silver-label table) plus 70 PDFs. This
module hands all of that to the pipeline as pandas DataFrames.
"""
from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import List

import pandas as pd
from openpyxl import load_workbook

from . import config


@dataclass(frozen=True)
class Row:
    """A single (Filename, Brand) target row from the Submissions sheet."""
    filename: str
    brand: str        # canonical form (e.g., "TREMFYA")
    pdf_path: Path


def load_submissions() -> List[Row]:
    """Return the 79 (Filename, Brand) target rows in order."""
    df = pd.read_excel(config.RULES_XLSX, sheet_name="Submissions")
    df = df[["Filename", "Brand"]].dropna(how="all")
    rows: List[Row] = []
    for _, r in df.iterrows():
        fn = str(r["Filename"]).strip()
        brand = config.canonical_brand(str(r["Brand"]))
        rows.append(Row(filename=fn, brand=brand, pdf_path=config.PDF_DIR / fn))
    return rows


def load_business_rules() -> pd.DataFrame:
    """The 12 parameter definitions, indexed by Param Name."""
    df = pd.read_excel(config.RULES_XLSX, sheet_name="Business Rules")
    df.columns = [str(c).strip() for c in df.columns]
    return df.set_index("Param Name")


def load_silver_labels() -> pd.DataFrame:
    """440 silver-label rows for OTHER brands/policies used for retrieval + calibration."""
    return pd.read_excel(config.RULES_XLSX, sheet_name="Additional Extracted Data")


def load_reference_example() -> dict:
    """Parse the single worked-example Reference sheet into a flat dict.

    The Reference sheet has columns: Sno., Params, Values, Reference Snips, Comments.
    Returns {param_name: {"value": ..., "comments": ...}}.
    """
    wb = load_workbook(config.RULES_XLSX, data_only=True)
    ws = wb["Reference"]
    out: dict = {}
    for r in range(5, ws.max_row + 1):
        param = ws.cell(row=r, column=6).value
        value = ws.cell(row=r, column=7).value
        comments = ws.cell(row=r, column=12).value
        if param:
            out[str(param).strip()] = {
                "value": value,
                "comments": (comments or "").strip() if isinstance(comments, str) else None,
            }
    return out


def list_pdfs() -> List[Path]:
    return sorted(p for p in config.PDF_DIR.glob("*.pdf"))


def summary() -> dict:
    rows = load_submissions()
    pdfs = list_pdfs()
    brand_counts: dict = {}
    multi_brand: set = set()
    seen: set = set()
    for r in rows:
        brand_counts[r.brand] = brand_counts.get(r.brand, 0) + 1
        if r.filename in seen:
            multi_brand.add(r.filename)
        seen.add(r.filename)
    return {
        "total_rows": len(rows),
        "total_pdfs": len(pdfs),
        "unique_filenames": len(seen),
        "multi_brand_pdfs": sorted(multi_brand),
        "brand_counts": brand_counts,
    }
